from openpyxl import load_workbook
wb2 = load_workbook('Move_2015_08_21_13_09_49_Running.xlsx')
print(wb2.get_sheet_names())
#['21 Aug 2015 13_09_49']
ws = wb2.active



#$seance_id=$_POST['seance_id'];
#$seance_name=$_POST['seance_name'];
#$sport_id=$_POST['sport_id'];
print(ws['A3'].value)
##1=velo
##2=escalade
##3=natation
##4=Footing
##5=Rando
##6=Ski de rando
##7=Skating
##8=SKI ALPIN
##10=musculation

#$date=$_POST['date'];
#$cal=$_POST['cal'];
print(ws['D3'].value)
#$dist=$_POST['dist'];
print(ws['E3'].value)
##in meter
#$duration=$_POST['duration'];
print(ws['C3'].value)
##in seconde to be converted HH:MM:ss

#$fat=$_POST['fat'];
#$above=$_POST['above'];
#$below=$_POST['below'];
#$in_zone=$_POST['in_zone'];
#$lower=$_POST['lower'];
#$upper=$_POST['upper'];

#$fmoy=$_POST['fmoy'];
print(ws['F3'].value)
#$fmax=$_POST['fmax'];
print(ws['S3'].value)
#$vmoy=$_POST['vmoy'];
print(ws['I3'].value)

#$vmax=$_POST['vmax'];
#$altitude=$_POST['altitude'];
#$url =  $_POST['url'];

#$action=$_POST['action'];

