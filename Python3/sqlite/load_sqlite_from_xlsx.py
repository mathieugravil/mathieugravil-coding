import csv
import sqlite3
from datetime import datetime
import re
import sys
from openpyxl import load_workbook

def create_table_from_xlsx(wb_name,ws_name,tablename, db_file, file_cmd):
  dll = "create table "+tablename+" ("
  wb = load_workbook(filename = wb_name)#, read_only=True)
  ws_servers = wb[ws_name]
  for i in range(1, ws_servers.get_highest_column()-1):
    dll = dll+re.sub('[^A-Za-z0-9]+','_',str(ws_servers.cell(row = 1, column = i).value))+" ,"
  dll = dll + re.sub('[^A-Za-z0-9]+','_',str(ws_servers.cell(row = 1, column = ws_servers.get_highest_column()-1).value))+" );"
  print(dll)
  conn = sqlite3.connect(db_file)
  c = conn.cursor()
  c.execute(str(dll))
  conn.commit()
  conn.close()


def insertfromxlsx(wb_name,ws_name,tablename,db_file,header,file_cmd):
  wb = load_workbook(filename = wb_name)#, read_only=True)
  ws_servers = wb[ws_name]
  i = 1
  fic = open (str(file_cmd), "a")
  cnx = sqlite3.connect(db_file)
  cursor = cnx.cursor()
  INSERT_CMD="INSERT INTO `"+tablename+"` VALUES (  "
 
  for row in ws_servers.rows:
    if ( i > header):
      j=1
      for cell in row:
        if( j < len(row)-1 ):
          INSERT_CMD=INSERT_CMD+" '"+re.sub('[^A-Za-z0-9\.;]+','',str(cell.value))+"' ,"
          j+=1
      INSERT_CMD=INSERT_CMD+" '"+re.sub('[^A-Za-z0-9\.;]+','',str(row[len(row)-1]))+"' );"
      fic.write(str(INSERT_CMD)+"\n")
      cursor.execute(str(INSERT_CMD))
      cnx.commit()
      INSERT_CMD="INSERT INTO `"+tablename+"` VALUES (  "
    i+=1
  fic.close
  cursor.close()
  cnx.close()

  


def create_table_from_csv(filename,tablename,db_file, file_cmd):
  spamReader = csv.reader(open(filename, newline='\n'), delimiter=';')
  fic = open (str(file_cmd), "a")
  i=0
  max_lenght= []
  typ = []
  label = []
  cnx = sqlite3.connect(db_file)
  cursor = cnx.cursor()
  for row in spamReader:
    if ( i == 0):
      for k in range(len(row)):
        if len(row[k]) != 0 :
          max_lenght.append(0)
        typ.append("int")
        label.append(row[k])
      print( "nb champ = "+str(len(max_lenght)))
    else:
      for j in range(len(max_lenght)):
        if len(row[j]) >  max_lenght[j] :
          max_lenght[j]=len(row[j])
        if ( typ[j] != "varchar" ):
          try:
            float(row[j])
          except ValueError:
            typ[j]="varchar"
          else:
            if ( typ[j] == "int" ):
              try:
                int(row[j])
              except ValueError:
                typ[j]="float"
              else:
                typ[j]="int"
          if  ( typ[j] == "varchar" ):
            try:
              datetime.strptime(row[j],'%d/%m/%Y')
            except:
              typ[j]="varchar"
            else:
              typ[j]="datetime"
    i+=1
  print("Nb ligne = "+str(i))
  DLL="CREATE TABLE "+tablename+" ("
  for l in range(len(max_lenght)-1):
    DLL=DLL+""+str(label[l])+" "+str(typ[l])+"("+str(max_lenght[l])+"),\n"
  DLL=DLL+""+str(label[len(max_lenght)-1])+" "+str(typ[len(max_lenght)-1])+"("+str(max_lenght[len(max_lenght)-1])+")\n"
  DLL=DLL+");"
 # DLL=DLL+" PRIMARY KEY (`id"+str(tablename)+"`));"
  print(str(DLL))
  cursor.execute(str(DLL))
  cnx.commit()
  cursor.close()
  cnx.close()
  fic.write(str(DLL)+"\n")
  fic.close

def insertfromcsv(db_file,tablename,filename,header,file_cmd):
  spamReader = csv.reader(open(filename, newline='\n'), delimiter=';')
  i=0
  fic = open (str(file_cmd), "a")
  cnx = sqlite3.connect(db_file)
  cursor = cnx.cursor()
  INSERT_CMD="INSERT INTO `"+tablename+"` VALUES (  "
  for row in spamReader:
    if ( i == 0 and header == "Y"):
      print( "On prend pas le header : nb champ = "+str(len(row)))
    else:
      for j in range(len(row)-1):
        INSERT_CMD=INSERT_CMD+" '"+str(row[j]).replace("'", r" " )+"' ,"
      INSERT_CMD=INSERT_CMD+" '"+str(row[len(row)-1]).replace("'", r" " )+"' );"
     # print(str(INSERT_CMD));
      fic.write(str(INSERT_CMD)+"\n")
      cursor.execute(str(INSERT_CMD))
      cnx.commit()
      INSERT_CMD="INSERT INTO `"+tablename+"` VALUES (  "
    i+=1
  fic.close
  cursor.close()
  cnx.close()


def main():
  try:
    input_file= sys.argv[1]
  except IndexError:
    print ("You must supply a csv file name.")
    sys.exit(2)
  try:
    db_file=sys.argv[2]
  except IndexError:
    print ("You must supply a db file name.")
    sys.exit(2)
  file_cmd="commad.sql"

  create_table_from_xlsx(input_file,'Servers','servers', db_file, file_cmd)
  insertfromxlsx(input_file,'Servers','servers',db_file,2,file_cmd)


  
  #create_table_from_csv(input_file,"DOC",db_file,file_cmd)
  #insertfromcsv(db_file,"DOC",input_file,"Y",file_cmd)


if __name__ == "__main__":
  main()
